import openpyxl as xl
import copy


import stream
from error import ClassError


class ClassMsgAddress:
    def __init__(self, root):
        self.Multi = False
        self.Union = root
        self.Intersection = []
        self.NowLine = {}


class ClassDataAddress:
    nEXTpATH = []
    nEXTcONL = []

    def __init__(self):
        self.Path = copy.deepcopy(ClassDataAddress.nEXTpATH[:-1])
        self.Conl = copy.deepcopy(ClassDataAddress.nEXTcONL[:-1])

        ClassDataAddress.nEXTpATH = []
        ClassDataAddress.nEXTcONL = []

        self.NowAddressData = stream.ClassShadowStream()

    @staticmethod
    def add(index, conl):
        ClassDataAddress.nEXTpATH += [index]
        ClassDataAddress.nEXTcONL += [conl]

    def finish(self):
        pass


class ClassXlsxDataCell:
    """处理中单个行数据表达

        数据处理中对每行单个数据的储存

    """
    def __init__(self):
        self.CountNextAppear = 1
        self.InputCount = 0
        self.Last = None
        self.DataAppear = stream.ClassShadowStream()    # 1 开

    def __getitem__(self, item):
        return self.DataAppear[item]

    def up_data(self, value):
        for key in self.DataAppear.keys():
            self.DataAppear[key] <<= 1

        self.InputCount += 1

        try:
            self.DataAppear[value] += 1
        except KeyError:
            self.DataAppear[value] = 1
            self.DataAppear.set(value, self.CountNextAppear)
            self.CountNextAppear += 1
            self.Last = value

        return self.DataAppear.value_(value)

    def get_value(self, count):
        index = [x for x in self.DataAppear.keys_()]
        try:
            if count >= 1:
                return index[count-1]
            else:
                return index[count]
        except IndexError:
            # 警告 引索超范围
            if count >= 1:
                return index[-1]
            else:
                return index[0]


class ClassXlsxData:
    """处理中数据表达

        对于每文件输入数据的累计表达, 由processor.xlsxProcessor 处理

    """
    def __init__(self, data: list):  # 仅初始化DataList
        self.DataList = stream.ClassStream()  # 用于储存ClassXlsxDataCell 1 开
        self.NowAppear = copy.deepcopy(data)                 # 0开
        self.CacheLine = copy.deepcopy(data)                # 0开
        self.Count = 0                        # 计数

        self.NowLine = copy.deepcopy(data)                   # 0开
        self.Multi = False
        self.AddressUnion = None
        self.AddressIntersection = None

        jian = 1
        for jj in data:
            self.DataList[jian] = ClassXlsxDataCell()
            jian += 1

    def __getitem__(self, item):
        return self.DataList[item + 1]

    def ask(self, column: int, enum) -> str or None:
        if not enum:
            return copy.deepcopy(self.CacheLine[column])

        value = self.NowLine[column]
        if not self.Multi:
            return value
        else:
            valueAddressCode = self.DataList[column+1][value]
            if not valueAddressCode & self.AddressUnion:
                self.NowLine[column] = None
                return None
            else:
                for code in self.AddressIntersection:
                    if not valueAddressCode & code:
                        self.NowLine[column] = None
                        return None
                return value

    def none(self, column: int):
        self.NowLine[column] = None

    def get_data(self, data: list):
        self.NowLine = copy.deepcopy(data)
        self.CacheLine = copy.deepcopy(data)
        self.Count <<= 1
        self.Count += 1

        for c in range(len(data)):
            self.NowAppear[c] = self.DataList[c + 1].up_data(data[c])

    def index(self, data_address: ClassDataAddress):
        self.fresh()

        msg = ClassMsgAddress(self.Count)

        index = ClassNode.tree_address(data_address.Path, data_address.Conl, self, msg)

        self.get_address_info(msg)

        return index

    def fresh(self):
        self.NowLine = copy.deepcopy(self.CacheLine)
        self.Multi = False
        self.AddressUnion = None
        self.AddressIntersection = None

    def get_address_info(self, msg: ClassMsgAddress):
        self.Multi = msg.Multi
        self.AddressUnion = msg.Union
        self.AddressIntersection = msg.Intersection

    def show(self):
        self.DataList[1].DataAppear.test_show()


class ClassNode:
    @staticmethod
    def tree_address(paths, conls, data: ClassXlsxData, msg: ClassMsgAddress):
        try:
            path = paths[0]
            conl = conls[0]

            jian = abs(conl)
            jin = data.DataList[path + 1].DataAppear.value_(data.CacheLine[path])
            jin = ((jin-1)//jian)*jian+1

            node = data.DataList[path + 1].DataAppear.value(data.CacheLine[path])
                
            if jian == 1:
                pass
            else:
                addS = [x for x in data.DataList[path + 1].values()]
                addS.reverse()
                part = 0
                if conl > 1:
                    while jian == 1:
                        part |= addS[jian - 1]
                        jian -= 1
                elif conl < -1:
                    msg.Multi = True
                    msg.Intersection = [node]
                    while jian == 1:
                        part |= addS[jian - 1]
                        msg.Intersection += [addS[jian - 1]]
                        jian -= 1
                else:
                    # error
                    pass
                node |= part

            msg.Union &= node
            return str(jin) + ClassNode.tree_address(paths[1:], conls[1:], data, msg)
        except IndexError:
            return ""


class ClassXlsxSource:
    def __init__(self, name, sheet=None):
        self.__XlsxBody = xl.load_workbook(name)
        if sheet is None:
            self.NowSheet = self.__XlsxBody.sheetnames[0]
        else:
            self.NowSheet = sheet

    def read_all_lines(self, start_row=1, start_column=1):
        if ClassError.TEST:
            count = 0
            sheet = self.__XlsxBody[self.NowSheet]

            cache = sheet.values

            for jian in cache:
                cache = list(jian)
                break

            long = len(cache)

            for row in sheet.iter_rows(min_row=start_row, min_col=start_column, values_only=True) :
                count += 1
                for it in range(long):
                    if row[it] is not None:
                        cache[it] = str(row[it])

                yield cache
                if count >= 21:
                    break
        else:
            sheet = self.__XlsxBody[self.NowSheet]

            cache = sheet.values

            for jian in cache:
                cache = list(jian)
                break

            long = len(cache)

            for row in sheet.iter_rows(min_row=start_row, min_col=start_column, values_only=True):

                for it in range(long):
                    if row[it] is not None:
                        cache[it] = str(row[it])

                yield cache


if __name__ == "__main__":
    namePath = "./data/test_irregular_table.xlsx"

    xlsx = ClassXlsxSource(namePath)

    for ii in xlsx.read_all_lines():
        for i in ii:
            print(i, end=' ')

        print('\n')
